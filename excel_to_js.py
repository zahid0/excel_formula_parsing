import argparse
import re

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple


def is_within_range(cell, min_cell, max_cell):
    """
    Checks if the cell (e.g., "B2") is within the range defined by [min_cell, max_cell].
    Coordinates are interpreted as (col, row) tuples.
    """
    col, row = coordinate_to_tuple(cell)
    min_col, min_row = coordinate_to_tuple(min_cell)
    max_col, max_row = coordinate_to_tuple(max_cell)

    return (min_col <= col <= max_col) and (min_row <= row <= max_row)


def sanitize_function_name(name):
    """
    Sanitize the input name to be used as a valid JavaScript function name.
    Rules:
      - Remove spaces and any invalid characters.
      - Ensure the first character is not a digit (prefix with "_" if needed).
    """
    # Remove any character that is not alphanumeric or underscore or $
    sanitized = re.sub(r"[^A-Za-z0-9_$]", "", name)
    # If empty, use a default name.
    if not sanitized:
        sanitized = "excelFunction"
    # If the first character is a number, prefix with an underscore.
    if re.match(r"^[0-9]", sanitized):
        sanitized = "_" + sanitized
    return sanitized


def excel_sheet_to_js_function(
    excel_filename, sheet_name=None, min_cell=None, max_cell=None
):
    """
    Reads an Excel (.xlsx) file and returns a string containing JavaScript code
    for a function and an example input object `d` containing only the necessary
    input cells used in formulas.
    """
    wb = load_workbook(excel_filename, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active
    js_func_name = sanitize_function_name(ws.title)
    cell_ref_pattern = r"(\b[A-Z]+[0-9]+\b)"

    computed_cells = []
    formula_dependencies = {}
    input_cells = {}

    # First pass: collect all computed cells, their dependencies, and input cells
    for row in ws.iter_rows():
        for cell in row:
            cell_coord = cell.coordinate
            if min_cell or max_cell:
                if not is_within_range(
                    cell_coord, min_cell or "A1", max_cell or "XFD1048576"
                ):
                    continue
            if cell.data_type == "f":
                formula = cell.value
                if isinstance(formula, str) and formula.startswith("="):
                    clean_formula = formula[1:]
                    referenced_cells = re.findall(cell_ref_pattern, clean_formula)
                    formula_dependencies[cell_coord] = referenced_cells
                    computed_cells.append((cell_coord, clean_formula))
            else:
                value = cell.value
                input_cells[cell_coord] = value

    # Collect only input cells that are referenced by formulas
    referenced_input_cells = set()
    for cell, refs in formula_dependencies.items():
        for ref in refs:
            if ref in input_cells:
                referenced_input_cells.add(ref)

    # Build dependency graph
    dependency_graph = {}
    for cell, _ in computed_cells:
        dependency_graph[cell] = []
        for ref_cell in formula_dependencies.get(cell, []):
            if any(c == ref_cell for c, _ in computed_cells):
                dependency_graph[cell].append(ref_cell)

    # Perform topological sort
    from collections import defaultdict, deque

    in_degree = defaultdict(int)
    graph = defaultdict(list)
    nodes = list(dependency_graph.keys())

    for node in nodes:
        for neighbor in dependency_graph[node]:
            if neighbor in nodes:
                graph[neighbor].append(node)
                in_degree[node] += 1

    queue = deque()
    for node in nodes:
        if in_degree.get(node, 0) == 0:
            queue.append(node)

    sorted_order = []
    while queue:
        node = queue.popleft()
        sorted_order.append(node)
        for neighbor in graph[node]:
            in_degree[neighbor] -= 1
            if in_degree[neighbor] == 0:
                queue.append(neighbor)

    if len(sorted_order) != len(nodes):
        raise ValueError("Circular dependency detected in formulas")

    cell_formula_map = {cell: formula for cell, formula in computed_cells}

    sorted_computed_cells = []
    for cell in sorted_order:
        formula = cell_formula_map[cell]

        def replace_reference(match):
            ref_cell = match.group(1)
            if ref_cell in cell_formula_map:
                return f'computed["{ref_cell}"]'
            else:
                return f'd["{ref_cell}"]'

        js_expression = re.sub(cell_ref_pattern, replace_reference, formula)
        sorted_computed_cells.append((cell, js_expression))

    # Build JavaScript function code
    js_lines = [
        f"function {js_func_name}(d) {{",
        "    // d is an object containing input cell values",
        "    let computed = {};",
    ]

    for cell_coord, expression in sorted_computed_cells:
        js_lines.append(f"    computed['{cell_coord}'] = {expression};")

    js_lines.append("    return computed;")
    js_lines.append("}")

    # Build the d object containing only referenced input cells
    d_lines = ["{"]
    for cell in sorted(referenced_input_cells):
        value = input_cells[cell]
        if isinstance(value, str):
            str_value = f'"{value}"'
        else:
            str_value = str(value)
        d_lines.append(f'    "{cell}": {str_value},')
    if referenced_input_cells:
        d_lines[-1] = d_lines[-1].rstrip(",")
    d_lines.append("};")
    d_object = "\n".join(d_lines)

    js_function_code = "\n".join(js_lines)
    return js_function_code, js_func_name, d_object


def main():
    # Setup argument parser
    parser = argparse.ArgumentParser(
        description="Convert Excel sheet to JavaScript function"
    )
    parser.add_argument("excel_file", type=str, help="Path to Excel file")
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Comma-separated list of sheet names to process. If not provided, all sheets will be processed.",
    )
    parser.add_argument(
        "--min_cell",
        type=str,
        default=None,
        help="Minimum cell coordinate to limit processing",
    )
    parser.add_argument(
        "--max_cell",
        type=str,
        default=None,
        help="Maximum cell coordinate to limit processing",
    )
    parser.add_argument(
        "--include-test-code",
        action="store_true",
        help="Include d_object and console.log test code in the output",
    )

    args = parser.parse_args()

    # Get sheet names
    wb = load_workbook(args.excel_file)

    if args.sheet:
        sheet_names = [s.strip() for s in args.sheet.split(",")]
    else:
        sheet_names = wb.sheetnames

    for sheet_name in sheet_names:
        try:
            js_function_code, function_name, d_object = excel_sheet_to_js_function(
                args.excel_file,
                sheet_name=sheet_name,
                min_cell=args.min_cell,
                max_cell=args.max_cell,
            )
            print(f"// Code for sheet: {sheet_name}")
            print(js_function_code)

            if args.include_test_code:
                print(
                    f'\n\nconst {function_name}_args = {d_object}\n\nconsole.log("Output for {function_name}: ", {function_name}({function_name}_args))\n'
                )

        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {str(e)}")


if __name__ == "__main__":
    main()
